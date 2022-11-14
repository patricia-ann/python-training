# Openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.
#
# References:
# https://openpyxl.readthedocs.io/en/stable/tutorial.html
# https://openpyxl.readthedocs.io/en/stable/editing_worksheets.html
# Install using pip install openpyxl

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from os import system
system("clear")



# Reading an Existing Workbook
wb = load_workbook('Python_Lessons/13_OpenPyxl/Grades.xlsx')

# Accessing active sheet in a Workbook, in this case it is "Grades" sheet
ws = wb.active

# To display the sheetnames in the workbook
print(wb.sheetnames)

# # Changing Sheet Name
# wb['Grades'].title = "Student Grades"

# # Saving Workbooks, need to close the workbook first
# wb.save('Python_Lessons/13_OpenPyxl/Grades.xlsx')

# Accessing Cell Value
print(ws['A2'].value)

# # Accessing a Range of Cells
# print(ws['A1:E1'])

# # Accessing Range of Rows and Columns
# col = ws['C:D']
# row = ws['1:2']

# print(row)
# print(col)

# # Using iter_rows to print values of cells per row
# for row in ws.iter_rows(min_row=1, max_col=5, max_row=2):
#     for cell in row:
#         print(cell.value)

# # Using iter_cols to print values of cells per column
# for col in ws.iter_cols(min_row=1, max_col=5, max_row=2):
#     for cell in col:
#         print(cell.value)

# # Creating a Sheet
# wb.create_sheet("test")

# # Removing a Sheet
# wb.remove(wb['test'])

# # Changing the value of cell
# ws['A1'] = 'Test'

# # Creating a New Workbook
# wb2 = Workbook()
# ws1 = wb2.active

# # Adding a sheet on new workbook
# ws1 = wb2.create_sheet("My_Sheet")

# # Adding/Appending Rows
# for row in range(1, 3):
#     ws1.append(['ELIT', 'Python', '101'])

# wb2.save('My_Workbook.xlsx')

# # Loading Newly Created Workbook
# wb2 = load_workbook('My_Workbook.xlsx')

# # Remove Default Sheet
# wb2.remove(wb2['Sheet'])

# # Merging Cells
# ws1 = wb2.active
# ws1.merge_cells('A3:C3')

# # Inserting and Deleting Rows
# ws1 = wb2['My_Sheet']
# ws1.delete_rows(2)
# ws1.delete_cols(3)
# wb2.save('My_Workbook.xlsx')

# # Practical Example, Formulas & Cell Styling
# # On the last row on Grades.xlsx workbook, we will put each of the average of grades per subject
# wb = load_workbook('Grades.xlsx')
# ws = wb.active

# # Get the row to put the average, in this case it is on row 7.
# rownum = ws.max_row + 1
# # Get the number of grades to be used on average formula, in this case it is 5.
# len_data = rownum - 2

# # Averaging the grades and displaying on last row
# # We will use range to select colums 2 to 6, we'll get the column number using get_column_letter on openpyxl
# # and then apply the formula on each of the columns and assign the values on cells of the last row

# for i in range(2, 6):
#     char = get_column_letter(i)
#     cellname = char + str(rownum)
#     ws[cellname] = f"=SUM({char + '2'}:{char + '6'})/{len_data}"

# # Using styles, display the average as bold and with blue font color
# for col in range(1, 6):
#     ws[get_column_letter(col) + '7'].font = Font(bold=True, color="0099CCFF")

# # Save the modified Grades workbook as NewGrades.xlsx
# wb.save("NewGrades.xlsx")
